#!/usr/bin/env python3
"""
generera_json.py
Läser Farjor_Sverige_v2_normaliserad.xlsx och skriver farjor_data.json.
Kör detta skript varje gång Excel-källfilen uppdateras.

Användning:  python3 generera_json.py
Output:      farjor_data.json (i samma mapp)
"""

import openpyxl, json, re, sys
from pathlib import Path
from datetime import date

from schedule_instances import build_base_instances
from verified_schema_overrides import apply_verified_schema_overrides

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH  = SCRIPT_DIR / "Farjor_Sverige_v2_normaliserad.xlsx"
JSON_OUT   = SCRIPT_DIR / "farjor_data.json"

if not XLSX_PATH.exists():
    print(f"FEL: Hittar inte {XLSX_PATH}")
    sys.exit(1)

def clean_time(s):
    if not s: return ''
    m = re.match(r'(\d{1,2}:\d{2})', str(s).strip())
    return m.group(1) if m else str(s).strip()

print(f"Läser {XLSX_PATH.name}...")
wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)

ws = wb['Schemaregister']
hdrs = [c.value for c in ws[1]]
schema = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(hdrs, row))
    if not d.get('Veckodag'): continue
    schema.append({
        "id":           len(schema) + 1,
        "kategori":     d.get('Kategori') or '',
        "rederi":       d.get('Rederi') or '',
        "rutt":         d.get('Rutt') or '',
        "avghamn":      d.get('Avgångshamn') or '',
        "ankhamn":      d.get('Ankomsthamn') or '',
        "veckodag":     d.get('Veckodag') or '',
        "avgtid":       clean_time(d.get('Avgångstid')),
        "avgtid_raw":   clean_time(d.get('Avgångstid')),
        "anktid":       str(d.get('Ankomsttid') or ''),
        "nasta_dag":    bool(d.get('AnkomstNästaDag')),
        "mot_sverige":  bool(d.get('MotSverige')),
        "fran_sverige": bool(d.get('FrånSverige')),
        "anmarkning":   d.get('Anmärkning') or '',
        "verifiering":  d.get('Verifiering') or '',
        "kalla":        d.get('Källa') or '',
    })

schema = apply_verified_schema_overrides(schema)

ws2 = wb['Schemaregister_Intervall']
hdrs2 = [c.value for c in ws2[1]]
intervall = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    d = dict(zip(hdrs2, row))
    if not d.get('Rederi'): continue
    intervall.append({
        "id":           len(intervall) + 1,
        "kategori":     d.get('Kategori') or '',
        "rederi":       d.get('Rederi') or '',
        "rutt":         d.get('Rutt') or '',
        "veckodagar":   d.get('Veckodagar (text)') or '',
        "avgtid":       d.get('Avgångstid (text)') or '',
        "anktid":       d.get('Ankomsttid (text)') or '',
        "mot_sverige":  bool(d.get('MotSverige')),
        "fran_sverige": bool(d.get('FrånSverige')),
        "anmarkning":   d.get('Anmärkning') or '',
        "frekvens":     d.get('Frekvens') or '',
        "verifiering":  d.get('Verifiering') or '',
        "kalla":        d.get('Källa') or '',
    })

rederier = sorted(set(r['rederi'] for r in schema if r['rederi']))
hamnar   = sorted(set(r['avghamn'] for r in schema if r['avghamn']) |
                  set(r['ankhamn'] for r in schema if r['ankhamn']))

out = {
    "meta": {
        "uppdaterad":      str(date.today()),
        "kalla":           XLSX_PATH.name,
        "schema_rader":    len(schema),
        "intervall_rader": len(intervall),
        "rederier":        rederier,
        "hamnar":          hamnar,
    },
    "schema":    schema,
    "intervall": intervall,
}

out["avgangsinstanser"] = build_base_instances(out)
out["meta"]["avgangsinstans_dagar"] = len(out["avgangsinstanser"])

JSON_OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')
kb = JSON_OUT.stat().st_size / 1024
print(f"OK  {JSON_OUT.name} klar -- {len(schema)} avgångar, {len(intervall)} intervallrutter ({kb:.0f} KB)")
print(f"    Uppdaterad: {date.today()}")
